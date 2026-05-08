from datetime import timedelta

import openpyxl
from dateutil.relativedelta import relativedelta
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.views.generic import TemplateView
from openpyxl.styles import Font, PatternFill

from employees.models import Employee
from incidents.models import Incident
from medical_checks.models import MedicalCheck
from trainings.models import Training, TrainingProgram, TrainingCategory
from trainings.requirements import is_program_required_for_employee
from trainings.services import check_employee_compliance


# Create your views here.
def overdue_trainings_report(request):
    overdue_date = timezone.now().date() + timedelta(days=30)
    employees = Employee.objects.filter(
        instructions__next_training_date__lte=overdue_date,
        is_active=True
    ).distinct()
    return render(request,
                  'reports/overdue_trainings.html',
                  {'employees': employees})


def reports_index(request):
    """Отображает главную страницу со ссылками на все отчеты."""
    return render(request, 'reports/reports_index.html')


def overdue_medical_checks_report(request):
    """Отчет по сотрудникам с приближающимся сроком медосмотра."""
    overdue_date = timezone.now().date() + timedelta(days=30)

    # Находим ID сотрудников, у которых есть предстоящий медосмотр
    employee_ids = MedicalCheck.objects.filter(
        next_check_date__lte=overdue_date,
        employee__is_active=True
    ).values_list('employee_id', flat=True)

    employees = Employee.objects.filter(
        id__in=set(employee_ids)).order_by('last_name')

    return render(request,
                  'reports/overdue_medical_checks.html',
                  {'employees': employees})


def incident_statistics_report(request):
    """Отчет со статистикой по типам инцидентов."""
    stats = Incident.objects.values('incident_type').annotate(
        count=Count('id')).order_by('incident_type')

    # Для отображения полных названий в шаблоне
    incident_type_map = dict(Incident.INCIDENT_TYPES)

    incident_stats = [
        {
            'type_name': incident_type_map.get(item['incident_type']),
            'count': item['count']
        }
        for item in stats
    ]

    return render(request, 'reports/incident_statistics.html',
                  {'incident_stats': incident_stats})


def training_plan_report(request):
    """
    Улучшенное представление для формирования списка направляемых на обучение
    с фильтрацией по категории обучения
    """
    programs = TrainingProgram.objects.all().select_related('category')
    categories = TrainingCategory.objects.filter(is_active=True)

    selected_program_id = request.GET.get('program')
    selected_category = request.GET.get('category')
    planning_horizon_months = int(request.GET.get('horizon', 6))

    employees_to_train = []
    selected_program = None
    today = timezone.now().date()
    horizon_date = today + relativedelta(months=planning_horizon_months)

    # Фильтрация программ
    if selected_program_id:
        selected_program = get_object_or_404(
            TrainingProgram, id=selected_program_id)
        programs = programs.filter(id=selected_program_id)
    elif selected_category:
        programs = programs.filter(category__code=selected_category)

    # Получаем всех активных сотрудников
    all_employees = Employee.objects.filter(
        is_active=True,
        termination_date__isnull=True
    ).select_related('position', 'department')

    # ────────────────────────────────────────────────────────
    # ОПТИМИЗАЦИЯ: получаем все последние обучения одним запросом
    # ────────────────────────────────────────────────────────

    # 1. Сначала получаем все записи обучения для нужных сотрудников и программ
    all_trainings = Training.objects.filter(
        employee__in=all_employees,
        program__in=programs
    ).select_related('employee', 'program').order_by('employee', 'program', '-training_date')

    # 2. Строим словарь: ключ = (employee_id, program_id) → последняя запись обучения
    trainings_dict = {}
    for training in all_trainings:
        key = (training.employee_id, training.program_id)
        # Так как записи отсортированы по убыванию даты,
        # первая встреченная для каждой пары и будет последней
        if key not in trainings_dict:
            trainings_dict[key] = training

    # ────────────────────────────────────────────────────────
    # Теперь проходим по сотрудникам и программам без доп. запросов
    # ────────────────────────────────────────────────────────
    for emp in all_employees:
        for program in programs:
            # Проверяем, нужна ли эта программа сотруднику
            if not is_program_required_for_employee(emp, program):
                continue

            # Получаем последнее обучение из предварительно собранного словаря
            last_training = trainings_dict.get((emp.id, program.id))

            needs_training = False
            reason = ""
            expiry_date = None

            if last_training:
                if program.frequency_months > 0:
                    expiry_date = last_training.training_date + relativedelta(
                        months=program.frequency_months
                    )
                    if expiry_date <= horizon_date:
                        needs_training = True
                        if expiry_date < today:
                            reason = f"Просрочено (истекло {expiry_date.strftime('%d.%m.%Y')})"
                        else:
                            reason = f"Истекает {expiry_date.strftime('%d.%m.%Y')}"
            else:
                needs_training = True
                reason = "Первичное обучение (не пройдено)"

            if needs_training:
                # Проверяем, нет ли уже записи для этого сотрудника и программы
                already_in_list = any(
                    e['employee'].id == emp.id and e['program'].id == program.id
                    for e in employees_to_train
                )
                if not already_in_list:
                    employees_to_train.append({
                        'employee': emp,
                        'program': program,
                        'reason': reason,
                        'expiry_date': expiry_date,
                        'last_training_date': last_training.training_date if last_training else None,
                        'priority': _calculate_priority(emp, program, expiry_date, today)
                    })

    # Сортировка по приоритету
    employees_to_train.sort(
        key=lambda x: (
            x['priority'],
            x['expiry_date'] or today))

    return render(request, 'reports/training_plan.html', {
        'programs': programs,
        'categories': categories,
        'employees': employees_to_train,
        'selected_program': selected_program,
        'selected_category': selected_category,
        'planning_horizon_months': planning_horizon_months,
    })


def export_training_plan_excel(request):
    """
    Экспорт плана обучения в Excel.
    Использует ту же логику, что и training_plan_report.
    """
    programs = TrainingProgram.objects.all().select_related('category')
    categories = TrainingCategory.objects.filter(is_active=True)

    selected_program_id = request.GET.get('program')
    selected_category = request.GET.get('category')
    planning_horizon_months = int(request.GET.get('horizon', 6))

    selected_program = None
    today = timezone.now().date()
    horizon_date = today + relativedelta(months=planning_horizon_months)

    if selected_program_id:
        selected_program = get_object_or_404(TrainingProgram, id=selected_program_id)
        programs = programs.filter(id=selected_program_id)
    elif selected_category:
        programs = programs.filter(category__code=selected_category)

    all_employees = Employee.objects.filter(
        is_active=True,
        termination_date__isnull=True
    ).select_related('position', 'department')

    # ── Сбор данных (та же оптимизированная логика) ──
    all_trainings = Training.objects.filter(
        employee__in=all_employees,
        program__in=programs
    ).select_related('employee', 'program').order_by('employee', 'program', '-training_date')

    trainings_dict = {}
    for training in all_trainings:
        key = (training.employee_id, training.program_id)
        if key not in trainings_dict:
            trainings_dict[key] = training

    employees_to_train = []
    for emp in all_employees:
        for program in programs:
            if not is_program_required_for_employee(emp, program):
                continue

            last_training = trainings_dict.get((emp.id, program.id))

            needs_training = False
            reason = ""
            expiry_date = None

            if last_training:
                if program.frequency_months > 0:
                    expiry_date = last_training.training_date + relativedelta(
                        months=program.frequency_months
                    )
                    if expiry_date <= horizon_date:
                        needs_training = True
                        if expiry_date < today:
                            reason = f"Просрочено (истекло {expiry_date.strftime('%d.%m.%Y')})"
                        else:
                            reason = f"Истекает {expiry_date.strftime('%d.%m.%Y')}"
            else:
                needs_training = True
                reason = "Первичное обучение (не пройдено)"

            if needs_training:
                already_in_list = any(
                    e['employee'].id == emp.id and e['program'].id == program.id
                    for e in employees_to_train
                )
                if not already_in_list:
                    employees_to_train.append({
                        'employee': emp,
                        'program': program,
                        'reason': reason,
                        'expiry_date': expiry_date,
                        'last_training_date': last_training.training_date if last_training else None,
                        'priority': _calculate_priority(emp, program, expiry_date, today)
                    })

    employees_to_train.sort(
        key=lambda x: (x['priority'], x['expiry_date'] or today))

    # ── Создание Excel файла ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "План обучения"

    # Заголовки
    headers = ['№', 'ФИО', 'Должность', 'Отдел', 'Программа', 'Категория',
               'Последнее обучение', 'Срок действия', 'Причина', 'Приоритет']
    ws.append(headers)

    # Стилизация заголовков
    header_fill = PatternFill(start_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Приоритеты словами
    priority_names = {
        1: 'Критический (просрочено)',
        2: 'Высокий (истекает через месяц)',
        3: 'Средний (истекает через 2 мес)',
        4: 'Низкий (истекает через 3 мес)',
        5: 'Плановый',
    }

    # Данные
    for idx, item in enumerate(employees_to_train, 1):
        emp = item['employee']
        program = item['program']
        ws.append([
            idx,
            f"{emp.last_name} {emp.first_name} {emp.middle_name}",
            str(emp.position) if emp.position else "-",
            str(emp.department) if emp.department else "-",
            program.name,
            program.category.name if program.category else "-",
            item['last_training_date'].strftime('%d.%m.%Y') if item['last_training_date'] else "Не пройдено",
            item['expiry_date'].strftime('%d.%m.%Y') if item['expiry_date'] else "-",
            item['reason'],
            priority_names.get(item['priority'], f"Приоритет {item['priority']}"),
        ])

    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Отправка файла
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'plan_obucheniya_{today.strftime("%d.%m.%Y")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _calculate_priority(employee, program, expiry_date, today):
    """
    Рассчитывает приоритет обучения (1 - highest, 5 - lowest)
    """
    if not expiry_date:
        return 2  # Первичное обучение - высокий приоритет

    days_until_expiry = (expiry_date - today).days

    if days_until_expiry < 0:
        return 1  # Просрочено - критический приоритет
    elif days_until_expiry <= 30:
        return 2  # Истекает в течение месяца
    elif days_until_expiry <= 60:
        return 3  # Истекает в течение 2 месяцев
    elif days_until_expiry <= 90:
        return 4  # Истекает в течение 3 месяцев
    else:
        return 5  # Плановое


def sout_report(request):
    """Отчет по специальной оценке условий труда"""
    from assessments.models import Workplace
    from django.utils import timezone

    today = timezone.now().date()
    workplaces = Workplace.objects.select_related(
        'sout', 'position', 'site').all()

    stats = {
        'total': workplaces.count(),
        'not_conducted': 0,
        'expired': 0,
        'warning': 0,
        'valid': 0,
    }

    for wp in workplaces:
        status = wp.sout_status
        if status in stats:
            stats[status] += 1

    return render(request, 'reports/sout_report.html', {
        'workplaces': workplaces,
        'stats': stats,
        'today': today,
    })


def documents_report(request):
    """Отчет по документам"""
    from documents.models import Document
    from django.utils import timezone

    today = timezone.now().date()
    documents = Document.objects.all().select_related('employee', 'category')

    stats = {
        'total': documents.count(),
        'overdue': documents.filter(end_date__lt=today).count(),
        'expiring_soon': documents.filter(
            end_date__gte=today,
            end_date__lte=today + timedelta(days=30)
        ).count(),
    }

    return render(request, 'reports/documents_report.html', {
        'documents': documents,
        'stats': stats,
    })


def risks_report(request):
    """Отчет по профессиональным рискам"""
    from assessments.models import RiskAssessment, Workplace

    workplaces = Workplace.objects.prefetch_related('risks').all()

    stats = {
        'total_workplaces': workplaces.count(),
        'with_risks': workplaces.filter(
            risks__isnull=False).distinct().count(),
        'high_risks': RiskAssessment.objects.filter(
            risk_level__in=[
                'high',
                'critical']).count(),
    }

    return render(request, 'reports/risks_report.html', {
        'workplaces': workplaces,
        'stats': stats,
    })


class ComplianceDashboardView(TemplateView):
    """Панель соответствия требованиям по обучению"""
    template_name = 'reports/compliance_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        all_employees = Employee.objects.filter(
            is_active=True,
            termination_date__isnull=True
        )

        violations_count = 0
        warnings_count = 0
        compliant_count = 0

        for emp in all_employees:
            status = check_employee_compliance(emp)
            has_violations = (
                status['expired_programs'] or
                status['expired_instructions']
            )
            has_warnings = (
                status['missing_programs'] or
                status['missing_instructions']
            )

            if has_violations:
                violations_count += 1
            elif has_warnings:
                warnings_count += 1
            else:
                compliant_count += 1

        context.update({
            'total_employees': all_employees.count(),
            'compliant_count': compliant_count,
            'violations_count': violations_count,
            'warnings_count': warnings_count,
            'compliance_rate': round(
                compliant_count / all_employees.count() * 100, 1
            ) if all_employees.count() > 0 else 0,
            'today': today,
        })

        return context


class ComplianceReportView(TemplateView):
    """Отчёт по соответствию конкретного сотрудника"""
    template_name = 'reports/compliance_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee = get_object_or_404(Employee, pk=kwargs['employee_pk'])
        compliance_status = check_employee_compliance(employee)

        context.update({
            'employee': employee,
            'compliance': compliance_status,
            'today': timezone.now().date(),
        })

        return context