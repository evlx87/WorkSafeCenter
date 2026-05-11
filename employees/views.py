import os
import shutil
import tempfile
import zipfile
from io import BytesIO
from io import StringIO

import openpyxl
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.management import call_command
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from openpyxl.styles import Font, PatternFill

from organization.models import Department
from trainings.services import check_employee_compliance
from .forms import EmployeeForm
from .models import Employee


# Create your views here.
class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'employees/employee_list.html'
    context_object_name = 'employees'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('position', 'department')
        search_query = self.request.GET.get('search_query', '')
        department_id = self.request.GET.get('department', '')

        if search_query:
            queryset = queryset.filter(last_name__icontains=search_query)

        if department_id:
            queryset = queryset.filter(department_id=department_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Передаем в шаблон все отделы для выпадающего списка
        context['departments'] = Department.objects.all()
        # Передаем текущие значения фильтров, чтобы форма "помнила" их
        context['search_query'] = self.request.GET.get('search_query', '')
        context['selected_department'] = self.request.GET.get('department', '')
        return context


class EmployeeDetailView(DetailView):
    model = Employee
    template_name = 'employees/employee_detail.html'
    context_object_name = 'employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Получаем статус соблюдения требований
        compliance_status = check_employee_compliance(self.object)
        context['compliance'] = compliance_status
        context['today'] = timezone.now().date()
        # Добавляем информацию о документах
        context['employee_documents'] = self.object.trainings.filter(
            document_type__isnull=False
        ).select_related('program')
        # Просто флаг, есть ли проблемы, для отображения блока предупреждения
        context['has_compliance_issues'] = any([
            compliance_status['missing_programs'],
            compliance_status['missing_instructions'],
            compliance_status['expired_programs'],
            compliance_status['expired_instructions']
        ])
        return context


class EmployeeCreateView(CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'employees/employee_form.html'
    success_url = reverse_lazy('employees:employee_list')


class EmployeeUpdateView(UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'employees/employee_form.html'
    success_url = reverse_lazy('employees:employee_list')


class EmployeeDeleteView(DeleteView):
    model = Employee
    template_name = 'employees/employee_confirm_delete.html'
    success_url = reverse_lazy('employees:employee_list')


@staff_member_required
def import_data(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        dry_run = request.POST.get('dry_run') == 'on'

        # Сохраняем Excel во временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            for chunk in excel_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        # Создаём временную директорию, куда распакуем сканы (если они есть)
        scans_tmp_dir = tempfile.mkdtemp()

        try:
            # Распаковка ZIP со сканами
            zip_file = request.FILES.get('zip_file')
            if zip_file:
                with zipfile.ZipFile(zip_file, 'r') as zf:
                    zf.extractall(scans_tmp_dir)

            # Перехват вывода команды
            out = StringIO()
            call_command(
                'import_employee_data',
                file_path=tmp_path,
                scans_dir=scans_tmp_dir,
                dry_run=dry_run,
                stdout=out,
                stderr=out,
            )
            message = out.getvalue()
            success = "ИМПОРТ ЗАВЕРШЁН" in message
        except Exception as e:
            message = f"Ошибка: {e}"
            success = False
        finally:
            os.unlink(tmp_path)  # удаляем временный Excel
            shutil.rmtree(scans_tmp_dir, ignore_errors=True)  # удаляем распакованные сканы

        return render(request, 'employees/import_result.html', {
            'message': message,
            'success': success,
            'dry_run': dry_run,
        })

    return render(request, 'employees/import_form.html')


@staff_member_required
def download_template(request):
    """Генерирует и отдаёт Excel-шаблон для заполнения данных."""
    wb = openpyxl.Workbook()

    # 0. Организация
    ws0 = wb.active
    ws0.title = "0. Организация"
    ws0.append([
        'Полное название организации', 'ИНН', 'КПП', 'ОГРН',
        'Юридический адрес', 'Контактный телефон (макс. 20 символов)'
    ])

    # 1. Площадки
    ws1 = wb.create_sheet("1. Площадки")
    ws1.append(['Название площадки', 'Адрес площадки', 'Ответственный за ОТ (ФИО)'])

    # 2. Подразделения
    ws2 = wb.create_sheet("2. Подразделения")
    ws2.append(['Название отдела', 'Описание', 'Вышестоящий отдел (Название)'])

    # 3. Должности
    ws3 = wb.create_sheet("3. Должности")
    ws3.append(['Название должности', 'Отдел', 'Описание'])

    # 4. Сотрудники
    ws4 = wb.create_sheet("4. Сотрудники")
    ws4.append([
        'Фамилия', 'Предыдущая фамилия', 'Имя', 'Отчество',
        'Должность', 'Отдел', 'Дата рождения (ГГГГ-ММ-ДД)',
        'Дата приема (ГГГГ-ММ-ДД)', 'Телефон', 'Email',
        'Руководитель (Да/Нет)', 'Педагогический работник (Да/Нет)',
        'Специалист по ОТ (Да/Нет)', 'Член комиссии по ОТ (Да/Нет)',
        'Председатель комиссии по ОТ (Да/Нет)', 'И.о. директора (Да/Нет)',
        'Освобожден от первичного инструктажа (Да/Нет)',
        'В декретном отпуске (Да/Нет)', 'Дата увольнения (ГГГГ-ММ-ДД)',
        'Номер приказа об увольнении'
    ])

    # 5. Программы обучения
    ws5 = wb.create_sheet("5. Программы")
    ws5.append([
        'Название программы',
        'Тип (SAFETY/FIRE/FIRST_AID/WORKING_HEIGHT/OTHER)',
        'Часы', 'Периодичность (мес)', 'Обязательна для всех (Да/Нет)'
    ])

    # 6. Обучение
    ws6 = wb.create_sheet("6. Обучение")
    ws6.append([
        'ФИО Сотрудника (Полностью, как в листе 4)',
        'Категория обучения (ОТ/ПБ/ЭБ/Первая помощь/Антитерроризм/Другое)',
        'Тип документа (Протокол/Удостоверение/Сертификат/Диплом/Без документа)',
        'Название программы в документе',
        'Дата обучения (ГГГГ-ММ-ДД)',
        'Номер документа',
        'Имя файла скана (ivanov_doc.pdf)'
    ])

    # Стилизация
    header_fill = PatternFill(start_color="D9EAD3", fill_type="solid")
    header_font = Font(bold=True)
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 25

    # Сохраняем в память
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=full_system_template.xlsx'
    return response
