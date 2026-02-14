import openpyxl
from django.core.management.base import BaseCommand
from openpyxl.styles import Font, PatternFill


class Command(BaseCommand):
    help = 'Создает корректный Excel-шаблон для заполнения всей БД'

    def handle(self, *args, **options):
        wb = openpyxl.Workbook()

        # 0. Реквизиты организации (Только поля, существующие в модели)
        ws0 = wb.active
        ws0.title = "0. Организация"
        ws0.append([
            'Полное название организации',
            'ИНН',
            'КПП',
            'ОГРН',
            'Юридический адрес',
            'Контактный телефон (макс. 20 символов)'
        ])

        # 1. Площадки (Site)
        ws1 = wb.create_sheet("1. Площадки")
        ws1.append([
            'Название площадки',
            'Адрес площадки',
            'Ответственный за ОТ (ФИО)'
        ])

        # 2. Подразделения (Department)
        ws2 = wb.create_sheet("2. Подразделения")
        ws2.append([
            'Название отдела',
            'Описание',
            'Вышестоящий отдел (Название)'
        ])

        # 3. Должности (Position)
        ws3 = wb.create_sheet("3. Должности")
        ws3.append([
            'Название должности',
            'Отдел',
            'Описание'
        ])

        # 4. Программы обучения (TrainingProgram)
        ws4 = wb.create_sheet("4. Программы")
        ws4.append([
            'Название программы',
            'Тип (SAFETY/FIRE/FIRST_AID/WORKING_HEIGHT/OTHER)',
            'Часы',
            'Периодичность (мес)',
            'Обязательна для всех (Да/Нет)'
        ])

        # 5. Сотрудники (Employee)
        ws5 = wb.create_sheet("5. Сотрудники")
        ws5.append([
            'Фамилия',
            'Имя',
            'Отчество',
            'Должность',
            'Отдел',
            'Дата рождения (ГГГГ-ММ-ДД)',
            'Дата приема (ГГГГ-ММ-ДД)',
            'Телефон',
            'Email',
            'Руководитель (Да/Нет)',
            'Педагогический работник (Да/Нет)',
            'Специалист по ОТ (Да/Нет)',
            'Член комиссии по ОТ (Да/Нет)',
            'Председатель комиссии по ОТ (Да/Нет)',
            'И.о. директора (Да/Нет)',
            'Освобожден от первичного инструктажа (Да/Нет)',
            'В декретном отпуске (Да/Нет)',
            'Дата увольнения (ГГГГ-ММ-ДД)',
            'Номер приказа об увольнении'
        ])

        # 6. Пройденное обучение (Training)
        ws6 = wb.create_sheet("6. Обучение")
        ws6.append([
            'ФИО Сотрудника (Полностью, как в листе 5)',
            'Название программы (как в листе 4)',
            'Дата обучения (ГГГГ-ММ-ДД)',
            'Номер протокола',
            'Номер удостоверения',
            'Имя файла скана (ivanov_doc.pdf)'
        ])

        # 7. Инструктажи (Instruction)
        ws7 = wb.create_sheet("7. Инструктажи")
        ws7.append([
            'ФИО Сотрудника (Полностью)',
            'Тип инструктажа (полное название из справочника)',
            'Дата проведения (ГГГГ-ММ-ДД)',
            'Инструктор / Ответственный',
            'Примечания'
        ])

        # Стилизация
        header_fill = PatternFill(start_color="D9EAD3", fill_type="solid")
        header_font = Font(bold=True)
        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col in sheet.columns:
                sheet.column_dimensions[col[0].column_letter].width = 25

        filename = 'full_system_template.xlsx'
        wb.save(filename)
        self.stdout.write(
            self.style.SUCCESS(f'✅ Создан корректный шаблон: {filename}')
        )
        self.stdout.write(
            self.style.WARNING(
                '⚠️ ВАЖНО: Поля "Директор" и "Спец. по ОТ" назначаются автоматически '
                'на основе флагов сотрудников (Руководитель / Специалист по ОТ) после импорта.'
            )
        )
