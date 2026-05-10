import os
from datetime import datetime, timedelta

import openpyxl
from django.core.files import File
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Q
from django.utils.dateparse import parse_date

from employees.models import Employee
from organization.models import OrganizationSafetyInfo, Site, Department, Position
from trainings.models import Training, TrainingProgram, TrainingCategory


class Command(BaseCommand):
    help = 'Импорт данных организации и сотрудников из Excel-файла (только из шаблона)'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Путь к Excel файлу')
        parser.add_argument(
            '--scans-dir',
            type=str,
            help='Папка с PDF-файлами удостоверений и сканов'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Режим проверки без сохранения изменений'
        )

    # ================================================================
    # Вспомогательные методы парсинга
    # ================================================================
    def _parse_date(self, value):
        """Безопасный парсинг даты из разных форматов Excel."""
        if not value:
            return None
        try:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, str):
                value = value.strip()
                if not value:
                    return None
                date_formats = [
                    '%Y-%m-%d', '%d.%m.%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%Y'
                ]
                for fmt in date_formats:
                    try:
                        return datetime.strptime(value, fmt).date()
                    except ValueError:
                        continue
                return parse_date(value)
            if isinstance(value, (int, float)):
                excel_epoch = datetime(1899, 12, 30)
                return (excel_epoch + timedelta(days=value)).date()
        except Exception as e:
            self.stdout.write(
                self.style.WARNING(
                    f'⚠️ Не удалось распарсить дату: {value} ({e})'))
            return None
        return None

    def _parse_boolean(self, value):
        """Безопасный парсинг булевых значений."""
        if not value:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return str(value).strip().lower() in (
                'да', 'yes', 'true', '1', '+')
        return bool(value)

    def _get_cell_value(self, row, index, default=None):
        """Безопасное получение значения ячейки."""
        try:
            if len(row) > index and row[index] is not None:
                value = row[index]
                if isinstance(value, str):
                    return value.strip() if value.strip() else default
                return value
            return default
        except Exception:
            return default

    # ================================================================
    # Строгий поиск сотрудника
    # ================================================================
    def _find_employee_strict(self, last_name, first_name, middle_name=''):
        """
        Ищет сотрудника строго по фамилии (включая предыдущую), имени и отчеству.
        При отсутствии вызывает CommandError.
        """
        qs = Employee.objects.filter(Q(last_name__iexact=last_name) | Q(
            previous_last_name__iexact=last_name), first_name__iexact=first_name)
        if middle_name:
            qs = qs.filter(middle_name__iexact=middle_name)
        emp = qs.first()
        if not emp:
            raise CommandError(
                f'Сотрудник "{last_name} {first_name} {middle_name}" не найден. '
                f'Убедитесь, что сотрудники импортированы перед обучением.'
            )
        return emp

    # ================================================================
    # Строгий поиск программы обучения
    # ================================================================
    def _get_program_or_raise(self, program_name, category_code):
        """
        Возвращает программу обучения строго по названию и категории.
        Если программа не найдена – CommandError.
        """
        program_name = str(program_name).strip()
        program = TrainingProgram.objects.filter(
            name__iexact=program_name,
            category__code=category_code
        ).first()
        if not program:
            raise CommandError(
                f'Программа "{program_name}" в категории {category_code} не найдена в справочнике.')
        return program

    # ================================================================
    # Основной обработчик
    # ================================================================
    def handle(self, *args, **options):
        path = options['file_path']
        scans_dir = options.get('scans-dir')
        dry_run = options.get('dry-run', False)

        if not os.path.exists(path):
            raise CommandError(f'❌ Файл не найден: {path}')

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            raise CommandError(f'❌ Ошибка открытия Excel файла: {e}')

        self.stdout.write(self.style.SUCCESS(f'✅ Файл загружен: {path}'))
        if dry_run:
            self.stdout.write(self.style.WARNING(
                '⚠️ РЕЖИМ ПРОВЕРКИ (dry-run) - изменения не сохраняются'))

        # Проверка директории сканов
        if scans_dir:
            if not os.path.isabs(scans_dir):
                scans_dir = os.path.abspath(scans_dir)

            if not os.path.exists(scans_dir):
                self.stdout.write(self.style.ERROR(
                    f'❌ Директория сканов не существует: {scans_dir}'))
                scans_dir = None
            elif not os.path.isdir(scans_dir):
                self.stdout.write(self.style.ERROR(
                    f'❌ Указанный путь не является директорией: {scans_dir}'))
                scans_dir = None
            elif not os.access(scans_dir, os.R_OK):
                self.stdout.write(self.style.ERROR(
                    f'❌ Нет прав на чтение директории: {scans_dir}'))
                scans_dir = None
            else:
                files_count = len([f for f in os.listdir(
                    scans_dir) if os.path.isfile(os.path.join(scans_dir, f))])
                self.stdout.write(
                    self.style.SUCCESS(
                        f'✅ Директория сканов доступна. Найдено файлов: {files_count}'))

        # Статистика
        stats = {
            'emp_created': 0,
            'emp_updated': 0,
            'train_created': 0,
            'train_errors': 0,
            'scan_loaded': 0,
            'scan_not_found': 0,
            'scan_error': 0,
        }

        # ====== 0. Организация ======
        org = None
        try:
            if "0. Организация" in wb.sheetnames:
                ws_org = wb["0. Организация"]
                for row in ws_org.iter_rows(min_row=2, values_only=True):
                    if self._get_cell_value(row, 0):
                        defaults = {
                            'name_full': str(self._get_cell_value(row, 0, '')).strip()[:255],
                            'inn': str(self._get_cell_value(row, 1, '')).strip()[:12],
                            'kpp': str(self._get_cell_value(row, 2, '')).strip()[:9],
                            'ogrn': str(self._get_cell_value(row, 3, '')).strip()[:15],
                            'address_legal': str(self._get_cell_value(row, 4, '')).strip()[:255],
                            'contact_phone': str(self._get_cell_value(row, 5, '')).strip()[:20],
                        }
                        if not dry_run:
                            org = OrganizationSafetyInfo.load_organization()
                            for key, value in defaults.items():
                                setattr(org, key, value)
                            org.save()
                        else:
                            org = type('Org', (), defaults)()
                        self.stdout.write(
                            self.style.SUCCESS(
                                f'✅ Организация: {
                                    defaults["name_full"]}'))
                        break
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(
                    f'❌ Ошибка импорта организации: {e}'))

        # ====== 1. Площадки ======
        try:
            if "1. Площадки" in wb.sheetnames:
                ws_sites = wb["1. Площадки"]
                for row in ws_sites.iter_rows(min_row=2, values_only=True):
                    if self._get_cell_value(row, 0):
                        if not dry_run:
                            Site.objects.get_or_create(
                                name=str(
                                    self._get_cell_value(
                                        row, 0, '')).strip(), organization=org, defaults={
                                    'address': str(
                                        self._get_cell_value(
                                            row, 1, '')).strip()[
                                        :255], 'ot_responsible_name': str(
                                        self._get_cell_value(
                                            row, 2, '')).strip()[
                                        :200], })
                self.stdout.write(
                    self.style.SUCCESS('✅ Площадки импортированы'))
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(
                    f'❌ Ошибка импорта площадок: {e}'))

        # ====== 2. Подразделения ======
        try:
            if "2. Подразделения" in wb.sheetnames:
                ws_deps = wb["2. Подразделения"]
                for row in ws_deps.iter_rows(min_row=2, values_only=True):
                    if self._get_cell_value(row, 0):
                        parent = None
                        if self._get_cell_value(row, 2):
                            parent = Department.objects.filter(
                                name=str(self._get_cell_value(row, 2)).strip()
                            ).first()
                        if not dry_run:
                            Department.objects.get_or_create(
                                name=str(
                                    self._get_cell_value(
                                        row,
                                        0,
                                        '')).strip(),
                                defaults={
                                    'description': str(
                                        self._get_cell_value(
                                            row,
                                            1,
                                            '')).strip(),
                                    'parent': parent})
                self.stdout.write(self.style.SUCCESS(
                    '✅ Подразделения импортированы'))
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(
                    f'❌ Ошибка импорта подразделений: {e}'))

        # ====== 3. Должности ======
        try:
            if "3. Должности" in wb.sheetnames:
                ws_pos = wb["3. Должности"]
                for row in ws_pos.iter_rows(min_row=2, values_only=True):
                    if self._get_cell_value(row, 0):
                        dept = None
                        if self._get_cell_value(row, 1):
                            dept = Department.objects.filter(
                                name=str(self._get_cell_value(row, 1)).strip()
                            ).first()
                        if not dry_run:
                            Position.objects.get_or_create(
                                name=str(
                                    self._get_cell_value(
                                        row, 0, '')).strip(), defaults={
                                    'department': dept, 'description': str(
                                        self._get_cell_value(
                                            row, 2, '')).strip()})
                self.stdout.write(
                    self.style.SUCCESS('✅ Должности импортированы'))
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(
                    f'❌ Ошибка импорта должностей: {e}'))

        # ====== 4. Сотрудники ======
        try:
            if "4. Сотрудники" not in wb.sheetnames:
                raise CommandError('Лист "4. Сотрудники" не найден!')

            ws_emp = wb["4. Сотрудники"]
            for row_idx, row in enumerate(
                    ws_emp.iter_rows(min_row=2, values_only=True), start=2):
                if not self._get_cell_value(
                        row,
                        0) or not self._get_cell_value(
                        row,
                        2):
                    continue
                try:
                    birth_date = self._parse_date(self._get_cell_value(row, 6))
                    hire_date = self._parse_date(self._get_cell_value(row, 7))
                    termination_date = self._parse_date(
                        self._get_cell_value(row, 18))

                    position = None
                    if self._get_cell_value(row, 4):
                        position = Position.objects.filter(
                            name=str(self._get_cell_value(row, 4)).strip()
                        ).first()

                    department = None
                    if self._get_cell_value(row, 5):
                        department = Department.objects.filter(
                            name=str(self._get_cell_value(row, 5)).strip()
                        ).first()

                    defaults = {
                        'middle_name': str(self._get_cell_value(row, 3, '')).strip()[:100],
                        'previous_last_name': str(self._get_cell_value(row, 1, '')).strip()[:100],
                        'position': position,
                        'department': department,
                        'birth_date': birth_date,
                        'hire_date': hire_date,
                        'phone': str(self._get_cell_value(row, 8, '')).strip()[:20],
                        'email': str(self._get_cell_value(row, 9, '')).strip()[:254],
                        'is_executive': self._parse_boolean(self._get_cell_value(row, 10)),
                        'is_pedagogical': self._parse_boolean(self._get_cell_value(row, 11)),
                        'is_safety_specialist': self._parse_boolean(self._get_cell_value(row, 12)),
                        'is_safety_committee_member': self._parse_boolean(
                            self._get_cell_value(row, 13)),
                        'is_safety_committee_chair': self._parse_boolean(
                            self._get_cell_value(row, 14)),
                        'is_acting_director': self._parse_boolean(self._get_cell_value(row, 15)),
                        'exempt_from_safety_instruction': self._parse_boolean(
                            self._get_cell_value(row, 16)),
                        'on_parental_leave': self._parse_boolean(self._get_cell_value(row, 17)),
                        'termination_date': termination_date,
                        'termination_order_number': str(
                            self._get_cell_value(row, 19, '')).strip()[:50],
                        'is_active': termination_date is None,
                    }

                    if not dry_run:
                        emp, created = Employee.objects.update_or_create(
                            last_name=str(
                                self._get_cell_value(
                                    row, 0, '')).strip()[
                                :100], first_name=str(
                                self._get_cell_value(
                                    row, 2, '')).strip()[
                                :100], defaults=defaults)
                        if created:
                            stats['emp_created'] += 1
                        else:
                            stats['emp_updated'] += 1
                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(
                            f'❌ Ошибка в строке {row_idx} (сотрудники): {e}'))
                    continue

            self.stdout.write(
                self.style.SUCCESS(
                    f'✅ Сотрудники: {stats["emp_created"]} новых, '
                    f'{stats["emp_updated"]} обновлено'))
        except Exception as e:
            raise CommandError(
                f'❌ Критическая ошибка при импорте сотрудников: {e}')

        # ====== 5. Программы обучения (пропускаем – справочник уже должен быть

        # ====== 6. Автоматическое назначение директора и специалиста по ОТ ===
        if org and not dry_run:
            try:
                director = Employee.objects.filter(
                    position__name__iexact='Директор', is_active=True
                ).order_by('hire_date').first()
                if not director:
                    director = Employee.objects.filter(
                        is_acting_director=True, is_active=True
                    ).order_by('hire_date').first()
                if director:
                    org.director = director
                    self.stdout.write(
                        self.style.SUCCESS(
                            f'✅ Назначен директор: {director}'))

                specialist = Employee.objects.filter(
                    is_safety_specialist=True, is_active=True
                ).first()
                if specialist:
                    org.safety_specialist = specialist
                    self.stdout.write(
                        self.style.SUCCESS(
                            f'✅ Назначен спец. по ОТ: {specialist}'))
                org.save()
            except Exception as e:
                self.stdout.write(
                    self.style.WARNING(
                        f'⚠️ Ошибка при назначении ответственных: {e}'))

        # ====== 7. Обучение (строгий импорт) ======
        category_map = {
            'от': 'SAFETY',
            'охрана труда': 'SAFETY',
            'пб': 'FIRE',
            'пожарная безопасность': 'FIRE',
            'эб': 'ELECTRICAL',
            'электробезопасность': 'ELECTRICAL',
            'первая помощь': 'FIRST_AID',
            'оказание первой помощи': 'FIRST_AID',
        }

        doc_type_map = {
            'протокол': 'PROTOCOL',
            'удостоверение': 'CERT_QUAL',
            'сертификат': 'CERT_COMPL',
            'диплом': 'DIPLOMA',
        }

        try:
            if "6. Обучение" in wb.sheetnames:
                ws_train = wb["6. Обучение"]
                for row_idx, row in enumerate(
                    ws_train.iter_rows(
                        min_row=2, values_only=True), start=2):
                    if not self._get_cell_value(
                            row,
                            0) or not self._get_cell_value(
                            row,
                            3):
                        continue

                    try:
                        fio_string = str(self._get_cell_value(row, 0)).strip()
                        fio_parts = fio_string.split()
                        if len(fio_parts) < 2:
                            raise CommandError(
                                f'Неверный формат ФИО в строке {row_idx}: "{fio_string}"')

                        last_name = fio_parts[0]
                        first_name = fio_parts[1]
                        middle_name = ' '.join(fio_parts[2:]) if len(
                            fio_parts) > 2 else ''

                        emp = self._find_employee_strict(
                            last_name, first_name, middle_name)

                        category_raw = str(
                            self._get_cell_value(
                                row, 1, 'другое')).strip().lower()
                        category_code = None
                        for key, code in category_map.items():
                            if key in category_raw:
                                category_code = code
                                break
                        if not category_code:
                            raise CommandError(
                                f'Неизвестная категория обучения "{category_raw}" ' f'в строке {row_idx}')

                        program_name_in_doc = str(
                            self._get_cell_value(row, 3, '')).strip()
                        train_date = self._parse_date(
                            self._get_cell_value(row, 4))
                        if not train_date:
                            raise CommandError(
                                f'Неверная дата обучения в строке {row_idx}')

                        doc_type_raw = str(
                            self._get_cell_value(
                                row, 2, '')).strip().lower()
                        doc_number = str(
                            self._get_cell_value(
                                row, 5, '')).strip()

                        program = self._get_program_or_raise(
                            program_name_in_doc, category_code)

                        if not dry_run:
                            # Создаём запись; если такая комбинация сотрудник-программа-дата
                            # уже существует, пропускаем (можно
                            # раскомментировать get_or_create)
                            _, created = Training.objects.get_or_create(
                                employee=emp,
                                program=program,
                                training_date=train_date,
                                defaults={
                                    'raw_program_name': program_name_in_doc[:500],
                                    'document_type': doc_type_map.get(doc_type_raw, 'OTHER'),
                                    'document_number': doc_number[:100] if doc_number else '',
                                }
                            )
                            if created:
                                stats['train_created'] += 1

                            # Загрузка скана
                            if scans_dir and self._get_cell_value(row, 6):
                                scan_file = str(
                                    self._get_cell_value(
                                        row, 6)).strip()
                                if not scan_file:
                                    self.stdout.write(
                                        self.style.WARNING(
                                            f'⚠️ Строка {row_idx}: имя файла скана пустое'))
                                    stats['scan_error'] += 1
                                else:
                                    file_path = os.path.join(
                                        scans_dir, scan_file)
                                    if os.path.exists(file_path) and os.access(
                                            file_path, os.R_OK):
                                        try:
                                            with open(file_path, 'rb') as f:
                                                training_obj = Training.objects.get(
                                                    employee=emp, program=program,
                                                    training_date=train_date)
                                                training_obj.document_scan.save(
                                                    scan_file, File(f), save=True)
                                                self.stdout.write(self.style.SUCCESS(
                                                    f'  ✅ Скан загружен: {scan_file}'))
                                                stats['scan_loaded'] += 1
                                        except Exception as e:
                                            self.stdout.write(self.style.ERROR(
                                                f'  ❌ Ошибка загрузки скана {scan_file}: {e}'))
                                            stats['scan_error'] += 1
                                    else:
                                        self.stdout.write(
                                            self.style.ERROR(
                                                f'  ❌ Файл скана не найден или нет прав: {file_path}'))
                                        stats['scan_not_found'] += 1

                    except CommandError as e:
                        self.stdout.write(
                            self.style.ERROR(
                                f'❌ Ошибка в строке {row_idx}: {e}'))
                        stats['train_errors'] += 1
                        continue

                self.stdout.write(
                    self.style.SUCCESS(
                        f'✅ Обучение: создано {
                            stats["train_created"]} записей'))
                if stats['train_errors'] > 0:
                    self.stdout.write(
                        self.style.ERROR(
                            f'❌ Ошибок в обучении: {
                                stats["train_errors"]}'))
            else:
                self.stdout.write(self.style.WARNING(
                    'ℹ️ Лист "6. Обучение" не найден'))
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(
                    f'❌ Ошибка при импорте обучения: {e}'))

        # ====== Итоговый отчёт ======
        self.stdout.write(self.style.SUCCESS('\n' + '=' * 60))
        self.stdout.write(self.style.SUCCESS('📊 ИТОГОВЫЙ ОТЧЁТ'))
        self.stdout.write(self.style.SUCCESS('=' * 60))
        self.stdout.write(
            f'✅ Сотрудники: {
                stats["emp_created"]} новых, {
                stats["emp_updated"]} обновлено')
        self.stdout.write(
            f'✅ Обучение: создано {
                stats["train_created"]} записей')
        if scans_dir:
            self.stdout.write(
                f'📎 Сканы: загружено {stats["scan_loaded"]}, '
                f'не найдено {stats["scan_not_found"]}, '
                f'ошибок {stats["scan_error"]}')
        if stats['train_errors'] > 0:
            self.stdout.write(
                self.style.ERROR(
                    f'❌ Ошибок обучения: {
                        stats["train_errors"]}'))
        self.stdout.write(self.style.SUCCESS('=' * 60))
        self.stdout.write(self.style.SUCCESS('ИМПОРТ ЗАВЕРШЁН'))

        if dry_run:
            self.stdout.write(self.style.WARNING(
                '⚠️ РЕЖИМ ПРОВЕРКИ - данные не сохранены'))
