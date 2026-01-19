from notifications.models import Notification
from .models import Workplace
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils import timezone


def check_sout_deadlines():
    """Проверяет сроки СОУТ и создает уведомления для ответственных."""
    workplaces = Workplace.objects.all()
    for wp in workplaces:
        status = wp.sout_status
        if status in ['not_conducted', 'expired', 'warning']:
            msg = f"Требуется проведение СОУТ для РМ №{wp.number} ({wp.position})"

            # Создаем уведомление, если такого еще нет (чтобы не спамить)
            Notification.objects.get_or_create(
                message=msg,
                defaults={'type': 'danger' if status != 'warning' else 'warning'}
            )


def export_sout_plan_to_excel():
    """Создает Excel-файл со списком РМ, требующих СОУТ."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "План СОУТ"

    # Заголовки
    headers = ['№ РМ', 'Площадка/Цех', 'Должность', 'Текущий статус', 'Срок следующей СОУТ']
    ws.append(headers)

    # Стилизация заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")

    # Получаем те же данные, что в SOUTPlanningListView
    today = timezone.now().date()
    # Мы можем импортировать логику фильтрации или просто пройти по всем
    workplaces = Workplace.objects.all()

    for wp in workplaces:
        status = wp.sout_status
        if status in ['not_conducted', 'expired', 'warning']:
            current = wp.get_current_sout()

            status_map = {
                'not_conducted': 'Не проводилась',
                'expired': 'Просрочена',
                'warning': 'Срок подходит'
            }

            ws.append([
                wp.number,
                str(wp.site) if wp.site else "—",
                str(wp.position),
                status_map.get(status),
                current.next_assessment_date.strftime('%d.%m.%Y') if current else "Требуется первичная"
            ])

    return wb